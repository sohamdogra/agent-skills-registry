#!/usr/bin/env python3
"""
LinkedIn Scout — Company People-Tab Reviewer
==============================================
Workflow (per company):
  1. Search for the company on LinkedIn
  2. Open the company page, jump straight to the "People" tab
  3. Auto-load more results + auto-highlight anyone already showing
     10+ mutual connections right there in the list (easy case)
  4. PAUSE — you manually review the page and note down:
       a) anyone with 10+ mutuals (already highlighted for you), OR
       b) anyone with <10 mutuals where one of THEIR mutuals has 20+
          mutuals with you (use the on-demand "check mutual" command
          below — LinkedIn doesn't expose this in the list view, so
          it can't be fully automated without opening every profile)
  5. You type a command to move to the next company, reload, load more
     people, check a specific mutual, or quit. Progress is saved so you
     can resume a half-finished run later.
"""

import asyncio
import json
import re
import sys
import time
import random
from typing import Optional
from pathlib import Path
from datetime import datetime

from playwright.async_api import async_playwright, Page, BrowserContext
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich import print as rprint

console = Console()

# ─────────────────────────────────────────────
# CONFIG — edit these to customize your search
# ─────────────────────────────────────────────

# ─────────────────────────────────────────────
# CONFIG — paste your company LinkedIn URLs here
# ─────────────────────────────────────────────
#
# Paste the company LinkedIn URLs to scan below, one per line.
# Any of these forms works (the script normalizes them to the People tab):
#   https://www.linkedin.com/company/kaiber/
#   https://www.linkedin.com/company/kaiber
#   linkedin.com/company/kaiber/people/
#   https://www.linkedin.com/company/kaiber/about/
#
COMPANY_URLS = [
    # Already processed earlier (kept so the progress file skips them):
    "https://www.linkedin.com/company/kaiber-ai/",
    "https://www.linkedin.com/company/colossyan/",
    "https://www.linkedin.com/company/hedra-labs/",
    "https://www.linkedin.com/company/videotto/",
    "https://www.linkedin.com/company/argilai/",
    "https://www.linkedin.com/company/creatify-ai/",
    "https://www.linkedin.com/company/arcads-ai/",
    "https://www.linkedin.com/company/avataros/",
    "https://www.linkedin.com/company/valka-ai/",
    "https://www.linkedin.com/company/decohere/",
    "https://www.linkedin.com/company/reactor-world/",
    "https://www.linkedin.com/company/haiperai/",
    "https://www.linkedin.com/company/genmoai/",
    "https://www.linkedin.com/company/moonvalley-ai/",
    "https://www.linkedin.com/company/craftstoryai/",
    "https://www.linkedin.com/company/wideframe/",
    "https://www.linkedin.com/company/mosaicai/",
    "https://www.linkedin.com/company/outpaint/",
    "https://www.linkedin.com/company/lemonslice/",
    "https://www.linkedin.com/company/yarn-so/",
    "https://www.linkedin.com/company/linum-ai/",
    "https://www.linkedin.com/company/tavus-io/",
    "https://www.linkedin.com/company/midrender/",
    "https://www.linkedin.com/company/kino-ai/",
    "https://www.linkedin.com/company/magichour/",
    "https://www.linkedin.com/company/latted/",
    "https://www.linkedin.com/company/focal-app/",
    "https://www.linkedin.com/company/diffusionstudio/",
    "https://www.linkedin.com/company/stewdioai/",
    "https://www.linkedin.com/company/palmierio/",
    "https://www.linkedin.com/company/knowlify/",
    "https://www.linkedin.com/company/bluma-ai/",
    "https://www.linkedin.com/company/koyal-ai/",
    "https://www.linkedin.com/company/absurd/",
    # New Companies that need to be processed
    
]

# Flag threshold for the "easy" criterion (visible directly on the People tab)
MIN_MUTUAL_CONNECTIONS = 10

# A mutual is considered "influential" (makes a <10 person worth noting) when
# they themselves share at least this many mutual connections with you.
MIN_MUTUAL_FRIEND_CONNECTIONS = 20

# How many mutuals to harvest/list per person (cap to keep runtime sane).
MAX_MUTUALS_PER_PERSON = 30

# When True, for each listed mutual we also open their profile to read how
# many mutuals THEY share with you (needed for the 20+ influential rule).
# This is slower and more profile-views, but you asked to check every person.
CHECK_EVERY_MUTUAL = True

# Keep all state/output in a STABLE per-user folder, not the current working
# directory and not the skill folder. Two reasons:
#  1. An agent (e.g. OpenClaw) may launch the script from any cwd; a fixed path
#     means the scan always finds the session login saved (no false
#     "not logged in").
#  2. Reinstalling/updating the skill wipes the skill directory — storing the
#     session there would destroy the login on every update. ~/.linkedin-scraper
#     survives updates.
# Override with the LINKEDIN_SCRAPER_HOME environment variable if desired.
import os as _os
_STATE_DIR = Path(_os.environ.get("LINKEDIN_SCRAPER_HOME", Path.home() / ".linkedin-scraper"))
_STATE_DIR.mkdir(parents=True, exist_ok=True)

SESSION_FILE = _STATE_DIR / "linkedin_session.json"
PROGRESS_FILE = _STATE_DIR / "scout_progress.json"
OUTPUT_XLSX = _STATE_DIR / "linkedin_scout_results.xlsx"
# Master cache of every company's scraped records across ALL runs. The
# spreadsheet is rebuilt from this each run, so previously-finished companies
# keep appearing even when they're skipped on a re-run.
RESULTS_FILE = _STATE_DIR / "scout_results.json"

# Human-like delays (seconds) — keep these; they're what keeps us under the
# bot-detection radar. Raise them if you start seeing checkpoints.
DELAY_SCROLL = (1, 2)
DELAY_AFTER_NAV = (2, 4)
DELAY_BETWEEN_COMPANIES = (4, 8)
DELAY_BETWEEN_PROFILES = (3, 6)   # pause between opening each person
DELAY_BETWEEN_MUTUALS = (2, 4)    # pause between opening each mutual

# ─────────────────────────────────────────────



async def human_delay(min_s: float = 1.0, max_s: float = 3.0):
    await asyncio.sleep(random.uniform(min_s, max_s))


async def async_input(prompt: str = "") -> str:
    """Non-blocking input() — runs the blocking call in a thread so
    Playwright's event loop keeps processing browser events (navigation,
    etc.) while we wait for the user to type something. Using plain input()
    here would freeze the event loop and cause page.url / clicks to look
    stale right after the prompt returns."""
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, input, prompt)


def normalize_company_people_url(url: str) -> Optional[str]:
    """Turn any company URL form into the canonical .../people/ URL.
    Accepts with/without scheme, trailing slash, or an existing sub-tab
    (/about/, /jobs/, etc.). Returns None if it isn't a company URL."""
    u = (url or "").strip()
    if not u:
        return None
    if not u.startswith("http"):
        u = "https://" + u.lstrip("/")
    # Grab the /company/<slug> portion
    m = re.search(r'(https?://[^/]*linkedin\.com/company/[^/?#]+)', u, re.IGNORECASE)
    if not m:
        return None
    base = m.group(1).rstrip("/")
    return base + "/people/"


def company_name_from_url(url: str) -> str:
    """Best-effort readable company name from the URL slug (used as a fallback
    label until we read the real name off the page)."""
    m = re.search(r'/company/([^/?#]+)', url or "", re.IGNORECASE)
    if not m:
        return url
    slug = m.group(1)
    return slug.replace("-", " ").replace("_", " ").title()


async def get_company_display_name(page: Page, fallback: str) -> str:
    """Read the company's real display name off the page (h1), falling back to
    the slug-derived name if not found."""
    try:
        name = await safe_eval(page, """() => {
            const h1 = document.querySelector('h1');
            return h1 ? (h1.innerText || '').trim() : '';
        }""")
        if name and len(name) < 100:
            return name
    except Exception:
        pass
    return fallback


async def current_url(page: Page) -> str:
    """Get the page's *actual* current URL via a live JS query, rather than
    Playwright's cached `page.url`, which can lag behind real navigation if
    the event loop was blocked (e.g. by a synchronous input() prompt)."""
    try:
        return await page.evaluate("window.location.href")
    except Exception:
        return page.url


async def safe_goto(page: Page, url: str, wait_until: str = "domcontentloaded", retries: int = 2) -> bool:
    """page.goto() with tolerance for LinkedIn's own redirects/interstitials
    (e.g. an account-switcher screen when a stale session cookie is present),
    which Playwright/Chromium often reports as net::ERR_ABORTED even though
    the browser ends up somewhere perfectly usable. Returns True if the goto
    call itself succeeded; False if it kept failing after retries (the
    browser may still be on a usable page even then — check current_url)."""
    for attempt in range(retries + 1):
        try:
            await page.goto(url, wait_until=wait_until)
            return True
        except Exception as e:
            if "ERR_ABORTED" in str(e) and attempt < retries:
                # Likely an in-flight redirect/interstitial — give it a beat and retry
                await human_delay(1.5, 2.5)
                continue
            if attempt < retries:
                await human_delay(1, 2)
                continue
            console.print(f"[dim red]  ⚠ Navigation hiccup going to {url}: {e}[/dim red]")
            return False
    return False


def parse_mutual_count(text: str):
    """Parse LinkedIn's real mutual-connection phrasings into an integer.
    Returns None if no mutual info is present at all.

    LinkedIn does NOT print 'N mutual connections' on People-tab / PYMK
    cards. It writes things like:
      'Ryan Paugh, Jerry Li, and 3 other mutual connections'  -> 2 + 3 = 5
      'Jerry Li, Ji Hong, and 11 other mutual connections'    -> 2 + 11 = 13
      'Jerry Li and 3 other mutual connections'               -> 1 + 3 = 4
      'Aury M. Cifuentes is a mutual connection'              -> 1
      'X and Y are mutual connections'                        -> 2
      '14 mutual connections'  (occasionally, e.g. tooltips)  -> 14

    NOTE: the card text often *wraps* so that 'mutual' and 'connections'
    land on different visual lines. We therefore collapse ALL whitespace
    (newlines included) into single spaces before matching — parsing
    line-by-line silently failed on wrapped text, which is what hid Ben
    Kusin and Xin Lu.
    """
    if not text or "mutual" not in text.lower():
        return None

    # Collapse newlines + runs of whitespace into single spaces
    flat = re.sub(r'\s+', ' ', text).strip()
    low = flat.lower()

    # Isolate the segment around 'mutual' so leading headline text (which may
    # itself contain commas) doesn't inflate the named-people count.
    seg_match = re.search(r'([^.;|]*?\bmutual connection[^.;|]*)', flat, re.IGNORECASE)
    segment = seg_match.group(1) if seg_match else flat
    seg_low = segment.lower()

    # "... and N other mutual connection(s)" → (named before) + N
    m = re.search(r'([^.;|]*?)\band\s+(\d+)\s+other\s+mutual\s+connection', segment, re.IGNORECASE)
    if m:
        named_part = m.group(1)
        n_others = int(m.group(2))
        named = [x for x in re.split(r',', named_part) if x.strip()]
        if not named and named_part.strip():
            named = [named_part.strip()]
        return len(named) + n_others

    # Bare "N mutual connection(s)" with no "other"
    m = re.search(r'(\d+)\s+mutual\s+connection', seg_low)
    if m:
        return int(m.group(1))

    # "X and Y are mutual connections" → 2 named, no "+N"
    if "are mutual connection" in seg_low:
        return 2

    # "X is a mutual connection" → 1
    if "is a mutual connection" in seg_low:
        return 1

    return None


def parse_degree(text: str) -> str:
    """Extract connection degree (1st / 2nd / 3rd) from card text."""
    m = re.search(r'\b(1st|2nd|3rd)\b', text, re.IGNORECASE)
    if m:
        return m.group(1).lower()
    # LinkedIn sometimes uses the superscript-style "· 2nd" or "2nd degree"
    m2 = re.search(r'(\d)(?:st|nd|rd)\s*(?:degree)?', text, re.IGNORECASE)
    if m2 and m2.group(1) in ("1", "2", "3"):
        return f"{m2.group(1)}{'st' if m2.group(1)=='1' else 'nd' if m2.group(1)=='2' else 'rd'}"
    return "unknown"


def is_non_person_name(name: str) -> bool:
    """True if a scanned 'name' is actually a non-person blurb LinkedIn renders
    in the People tab (service-provider cards, 'works here'/'follows this page'
    rows, etc.) — these must never be treated as people."""
    if not name:
        return True
    low = name.strip().lower()
    junk_prefixes = (
        "provides services",
        "provides services -",
    )
    junk_contains = (
        "works here",
        "follows this page",
        "follow this page",
        "people you may know",
        "show all",
        "see all",
    )
    if any(low.startswith(p) for p in junk_prefixes):
        return True
    if any(k in low for k in junk_contains):
        return True
    return False


async def safe_eval(page: Page, script: str, arg=None):
    """page.evaluate() that tolerates the page navigating out from under us —
    e.g. the user clicks 'back' in the browser while the script is mid-scroll.
    Returns None instead of raising 'Execution context was destroyed'."""
    try:
        if arg is not None:
            return await page.evaluate(script, arg)
        return await page.evaluate(script)
    except Exception as e:
        msg = str(e)
        if any(k in msg for k in ("Execution context was destroyed",
                                  "Target closed", "navigation", "detached")):
            return None
        return None


async def scroll_page(page: Page, times: int = 3):
    for _ in range(times):
        result = await safe_eval(page, "window.scrollBy(0, window.innerHeight * 0.8)")
        if result is None:
            # Page likely navigated (user interaction) — stop scrolling quietly
            break
        await human_delay(*DELAY_SCROLL)


async def thorough_scroll(page: Page, max_steps: int = 12):
    """Scroll the full page in steps (and back up) to force LinkedIn to
    lazy-render sections like 'People you may know', which only mount into
    the DOM once scrolled near the viewport. Without this, a scan can find
    zero profile cards even though they're visually 'there' after you scroll
    manually."""
    try:
        last_height = await safe_eval(page, "document.body.scrollHeight")
        if last_height is None:
            return
        for _ in range(max_steps):
            if await safe_eval(page, "window.scrollBy(0, window.innerHeight)") is None:
                return
            await human_delay(0.8, 1.6)
            new_height = await safe_eval(page, "document.body.scrollHeight")
            at_bottom = await safe_eval(
                page,
                "(window.innerHeight + window.scrollY) >= (document.body.scrollHeight - 200)"
            )
            if new_height is None:
                return
            if at_bottom and new_height == last_height:
                break
            last_height = new_height
        await safe_eval(page, "window.scrollTo(0, 0)")
        await human_delay(0.5, 1.0)
    except Exception:
        pass


async def click_show_more(page: Page, max_clicks: int = 5) -> int:
    """Click LinkedIn's 'Show more results' button on the People tab to load
    more cards. Returns how many times it actually clicked."""
    clicks = 0
    for _ in range(max_clicks):
        await scroll_page(page, 1)
        btn = await page.query_selector(
            "button:has-text('Show more results'), button:has-text('Load more')"
        )
        if not btn:
            break
        try:
            await btn.click()
            clicks += 1
            await human_delay(1.5, 3)
        except Exception:
            break
    return clicks


def load_progress() -> set:
    if PROGRESS_FILE.exists():
        try:
            return set(json.loads(PROGRESS_FILE.read_text()).get("done", []))
        except Exception:
            return set()
    return set()


def save_progress(done: set):
    PROGRESS_FILE.write_text(json.dumps({"done": sorted(done)}, indent=2))


def load_results() -> list:
    """Load the master cache of previously-scraped company results so the
    spreadsheet can include companies that are skipped on this run."""
    if RESULTS_FILE.exists():
        try:
            data = json.loads(RESULTS_FILE.read_text(encoding="utf-8"))
            return data if isinstance(data, list) else []
        except Exception:
            return []
    return []


def save_results(all_results: list):
    """Persist the full set of company results so re-runs don't lose data."""
    try:
        RESULTS_FILE.write_text(json.dumps(all_results, indent=2), encoding="utf-8")
    except Exception as e:
        console.print(f"[yellow]  ⚠ Couldn't save results cache: {e}[/yellow]")


async def save_session(context: BrowserContext):
    storage = await context.storage_state()
    SESSION_FILE.write_text(json.dumps(storage, indent=2))
    console.print(f"[green]✓ Session saved to {SESSION_FILE}[/green]")


async def login_linkedin(page: Page, context: BrowserContext):
    console.print("\n[bold yellow]📧 LinkedIn Login Required[/bold yellow]")
    console.print("A browser window will open. Please log in manually, then come back here.")

    ok = await safe_goto(page, "https://www.linkedin.com/login")
    if not ok:
        # Most likely an account-switcher interstitial intercepted the
        # navigation. Give the browser a moment to settle and keep going
        # rather than crashing — the user can still log in manually.
        await human_delay(2, 3)

    try:
        await page.wait_for_load_state("networkidle", timeout=8000)
    except Exception:
        pass

    url_now = await current_url(page)

    if "feed" in url_now or "mynetwork" in url_now:
        console.print("[green]✓ Already logged in![/green]")
        return

    if any(s in url_now for s in ("checkpoint", "challenge")):
        console.print(
            "[yellow]⚠ LinkedIn is showing a checkpoint/challenge screen "
            "(common when switching accounts). Resolve it manually in the browser.[/yellow]"
        )

    console.print(
        "[dim]If you see a screen offering to continue as your previous account, choose "
        "'Sign in with a different account' (or log out first) to switch — otherwise it'll "
        "just re-use the old session.[/dim]"
    )
    console.print("[dim]Waiting for you to log in in the browser window...[/dim]")
    await async_input(">>> Press Enter after logging in: ")

    try:
        await page.wait_for_url("**/feed/**", timeout=15000)
    except Exception:
        url_now = await current_url(page)
        if "feed" not in url_now and "mynetwork" not in url_now:
            console.print(
                f"[yellow]⚠ Still not seeing the feed (currently at: {url_now}). "
                f"Continuing anyway — make sure you're actually logged in before proceeding.[/yellow]"
            )

    console.print("[green]✓ Login successful![/green]")
    await save_session(context)


async def open_company_search_and_wait(page: Page, context: BrowserContext, company: str) -> tuple[str, Optional[Page], Optional[str]]:
    """Open LinkedIn's company search for `company` and let the user manually
    click into the right one — automatic detection is unreliable when
    multiple companies share a name or LinkedIn's markup shifts. Waits until
    the user confirms they're on the People tab, or chooses to skip/quit.
    Returns (status, active_page, people_url) where status is
    'ok' | 'skip' | 'quit'. active_page may be a DIFFERENT tab than `page`
    if LinkedIn opened the company in a new tab."""

    # Close leftover company tabs from previous companies so the next pick
    # can't be confused with a stale one, then run the search in our tab.
    await close_stale_company_tabs(context, keep=page)
    try:
        await page.bring_to_front()
    except Exception:
        pass

    search_url = f"https://www.linkedin.com/search/results/companies/?keywords={company.replace(' ', '%20')}"
    await safe_goto(page, search_url)
    await human_delay(*DELAY_AFTER_NAV)

    console.print(Panel(
        f"Search results for '[bold]{company}[/bold]' are open in the browser window.\n"
        f"1. Click on the correct company\n"
        f"2. Click its [bold]People[/bold] tab\n"
        f"3. Come back here and press Enter",
        title="👉 Pick the right company manually",
        border_style="yellow",
    ))

    while True:
        cmd = (await async_input(
            ">>> [Enter=I'm on the People tab, s=skip this company, q=quit] Command: "
        )).strip().lower()

        if cmd == "q":
            return "quit", None, None
        if cmd == "s":
            return "skip", None, None

        # Look across ALL open tabs — the company may have opened in a new one
        active_page, url_now = await find_company_page(context, page)

        if "/company/" in url_now and "/people" in url_now:
            return "ok", active_page, url_now

        if "/company/" in url_now:
            # On the company page but not the People tab yet — try clicking it for them
            people_tab = await active_page.query_selector("a[href*='/people/']")
            if people_tab:
                await people_tab.click()
                await human_delay(*DELAY_AFTER_NAV)
                url_now = await current_url(active_page)
                if "/people" in url_now:
                    return "ok", active_page, url_now
            console.print(
                f"[yellow]You're on the company page ({url_now}) but not the People tab yet — "
                f"click 'People' in the browser, then press Enter again here.[/yellow]"
            )
        else:
            console.print(
                f"[yellow]Currently showing: {url_now}\n"
                f"Doesn't look like you've clicked into a company page yet (checked all "
                f"{len(context.pages)} open tab(s)) — click the right company in the browser, "
                f"then press Enter again here.[/yellow]"
            )


async def find_company_page(context: BrowserContext, fallback: Page) -> tuple[Page, str]:
    """LinkedIn frequently opens a clicked result in a NEW browser tab, so the
    Page object the script holds can end up pointing at the old/empty search
    tab. Scan every open tab and return the one that's actually on a /company/
    URL (preferring a /people view, and the MOST RECENTLY opened such tab,
    since older company tabs may linger from previous companies)."""
    company_people = []  # (index, page, url)
    company_any = []
    try:
        for i, p in enumerate(context.pages):
            try:
                u = await p.evaluate("window.location.href")
            except Exception:
                u = p.url
            if "/company/" in u:
                if "/people" in u:
                    company_people.append((i, p, u))
                else:
                    company_any.append((i, p, u))
    except Exception:
        pass

    # Most recently opened tab = highest index in context.pages
    if company_people:
        _, p, u = max(company_people, key=lambda t: t[0])
        return p, u
    if company_any:
        _, p, u = max(company_any, key=lambda t: t[0])
        return p, u

    try:
        return fallback, await fallback.evaluate("window.location.href")
    except Exception:
        return fallback, fallback.url


async def close_stale_company_tabs(context: BrowserContext, keep: Page):
    """Close any leftover /company/ tabs from previous companies so the next
    company's tab can't be confused with an old one. Never closes `keep`
    (the script's primary search tab) or the last remaining tab."""
    try:
        for p in list(context.pages):
            if p is keep:
                continue
            try:
                u = await p.evaluate("window.location.href")
            except Exception:
                u = p.url
            if "/company/" in u and len(context.pages) > 1:
                try:
                    await p.close()
                except Exception:
                    pass
    except Exception:
        pass


async def detect_block(page: Page) -> Optional[str]:
    """Check whether LinkedIn is showing an auth wall, security checkpoint,
    rate-limit, or 'sign in to continue' interstitial instead of real
    content. Returns a human-readable reason string if blocked, else None."""
    try:
        url = await current_url(page)
        url_low = url.lower()
        if any(s in url_low for s in ("authwall", "/checkpoint", "/uas/login", "/login")):
            return f"LinkedIn redirected to a login/checkpoint URL ({url})"

        body = (await page.evaluate("document.body.innerText") or "").lower()
        block_phrases = [
            "sign in to continue",
            "join linkedin",
            "please verify",
            "unusual activity",
            "we restricted your account",
            "let’s do a quick security check",
            "let's do a quick security check",
            "you’ve reached the weekly limit",
            "you've reached the weekly limit",
            "to continue, sign in",
        ]
        for p in block_phrases:
            if p in body:
                return f"Page text suggests a block/auth wall: '{p}'"
    except Exception:
        pass
    return None


async def dump_debug(page: Page, label: str = "scan") -> Path:
    """Save a screenshot, the full HTML, the page innerText, and every anchor
    (href + text) to a timestamped folder. This lets us see EXACTLY what the
    script is looking at when a scan finds nothing, instead of guessing at
    LinkedIn's DOM."""
    debug_dir = Path("debug")
    debug_dir.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = debug_dir / f"{label}_{stamp}"

    try:
        # Quick viewport shot; if fonts hang we bail fast — the HTML/JSON
        # dump below is the genuinely useful diagnostic anyway.
        await page.screenshot(path=str(base.with_suffix(".png")), timeout=2500)
    except Exception:
        console.print("[dim red]  (screenshot skipped — HTML/JSON dump still saved)[/dim red]")

    try:
        html = await page.content()
        base.with_suffix(".html").write_text(html, encoding="utf-8")
    except Exception:
        pass

    try:
        info = await page.evaluate(
            """() => {
                const anchors = Array.from(document.querySelectorAll('a')).map(a => ({
                    href: a.getAttribute('href') || '',
                    text: (a.innerText || '').trim().slice(0, 80)
                }));
                const inLinks = anchors.filter(a => (a.href||'').includes('/in/'));
                const mutualMentions = (document.body.innerText.match(/mutual connection/gi) || []).length;
                return {
                    url: location.href,
                    anchor_total: anchors.length,
                    in_links: inLinks,
                    in_link_count: inLinks.length,
                    mutual_mentions: mutualMentions,
                    body_text: (document.body.innerText || '').slice(0, 4000),
                };
            }"""
        )
        base.with_suffix(".json").write_text(json.dumps(info, indent=2), encoding="utf-8")
        console.print(
            f"[yellow]  🩺 Debug dump saved to {base}.* — "
            f"{info.get('anchor_total', 0)} total links, "
            f"{info.get('in_link_count', 0)} profile (/in/) links, "
            f"{info.get('mutual_mentions', 0)} 'mutual connection' mentions on page.[/yellow]"
        )
        return base
    except Exception as e:
        console.print(f"[dim red]  (couldn't gather debug info: {e})[/dim red]")
    return base


async def scan_visible_people_hints(page: Page, retries: int = 4, reload_url: Optional[str] = None) -> list[dict]:
    """Scan for people, retrying a few times if the first pass finds nothing.
    LinkedIn lazy-loads the People-tab cards, so the very first scan right
    after navigation can land before the cards mount. We re-scroll, click
    'Show more', and (if given reload_url) re-navigate fresh between attempts,
    since a fresh navigation is what reliably forces the list to render."""
    hints = []
    for attempt in range(retries):
        hints = await _scan_once(page, dump_on_empty=(attempt == retries - 1))
        if hints:
            if attempt > 0:
                console.print(f"[dim]  (found people on retry {attempt + 1})[/dim]")
            return hints
        if attempt < retries - 1:
            console.print(f"[dim]  (nothing yet — reloading & waiting for cards, retry "
                          f"{attempt + 2}/{retries}...)[/dim]")
            # A fresh navigation is the most reliable way to force the People
            # list to mount; fall back to scroll/show-more if no URL given.
            if reload_url:
                await safe_goto(page, reload_url)
                await human_delay(2.0, 3.0)
            else:
                await human_delay(3.0 + attempt * 1.5, 4.5 + attempt * 1.5)
            await click_show_more(page, max_clicks=2)
            await thorough_scroll(page)
    return hints


async def _scan_once(page: Page, dump_on_empty: bool = True) -> list[dict]:
    """Extract every profile (anchor to /in/) currently in the DOM, along
    with the surrounding card text, in a single in-browser pass. Doing the
    walk inside page.evaluate is far more reliable than per-element Playwright
    queries, which can miss virtualized / lazily-mounted cards. We scroll the
    whole page first so lazy sections ('People you may know', associated
    members) actually render before we read them."""

    # Force lazy content to mount
    await thorough_scroll(page)

    # Is LinkedIn showing a wall instead of content?
    block_reason = await detect_block(page)
    if block_reason:
        console.print(f"[bold red]🚫 {block_reason}[/bold red]")
        console.print("[yellow]This usually means you're logged out, hit a security check, or are being "
                      "rate-limited. Resolve it in the browser window, then re-scan.[/yellow]")
        await dump_debug(page, "blocked")
        return []

    # One JS pass: collect each unique /in/ link + best surrounding text block.
    # Scope to the MAIN content column only — LinkedIn's right rail ('People
    # also viewed', promoted) and footer also contain /in/ links for people
    # who aren't on the People tab (this is how 'Ben Slater' leaked in).
    raw = await page.evaluate(
        """() => {
            const out = [];
            const seen = new Set();

            // Prefer the main content region; fall back to whole document.
            let root = document.querySelector('main')
                    || document.querySelector('.scaffold-layout__main')
                    || document.querySelector('[role=\"main\"]')
                    || document;

            // Elements we never want to read people from (sidebars/ads/footer)
            const isInExcludedRegion = (el) => {
                let n = el;
                for (let i = 0; i < 12 && n; i++) {
                    const cls = (n.className && n.className.toString) ? n.className.toString() : '';
                    const aside = n.tagName === 'ASIDE' || n.tagName === 'FOOTER';
                    if (aside) return true;
                    if (/aside|right-rail|secondary|ad-banner|footer|promo|people-also|similar/i.test(cls)) return true;
                    n = n.parentElement;
                }
                return false;
            };

            const anchors = Array.from(root.querySelectorAll("a[href*='/in/']"));
            for (const a of anchors) {
                if (isInExcludedRegion(a)) continue;
                let href = a.getAttribute('href') || '';
                if (!href) continue;
                try { href = new URL(href, location.origin).pathname; } catch (e) {}
                if (!href.includes('/in/')) continue;
                if (seen.has(href)) continue;
                seen.add(href);

                // The anchor's OWN text is the most reliable name for THIS
                // person — using the surrounding block's first line caused
                // names to be borrowed from neighbouring cards.
                let anchorName = (a.innerText || '').trim().split('\\n')[0].trim();
                // Some cards put the name in an aria-label or a child span
                if (!anchorName) {
                    anchorName = (a.getAttribute('aria-label') || '').trim();
                }
                if (!anchorName) {
                    const span = a.querySelector('span[aria-hidden=\"true\"], span');
                    if (span) anchorName = (span.innerText || '').trim().split('\\n')[0].trim();
                }

                // Climb for the mutual-connection TEXT only (count), bounded
                // so we don't swallow neighbouring cards.
                let node = a;
                let best = (a.innerText || '').trim();
                for (let i = 0; i < 5 && node; i++) {
                    node = node.parentElement;
                    if (!node) break;
                    const t = (node.innerText || '').trim();
                    if (/mutual/i.test(t) && t.length < 900) { best = t; break; }
                    if (t.length > best.length && t.length < 900) best = t;
                }
                out.push({ href, text: best, anchor_name: anchorName });
            }
            const mutualMentions = (document.body.innerText.match(/mutual connection/gi) || []).length;
            return { total_anchors: anchors.length, people: out, mutual_mentions: mutualMentions };
        }"""
    )

    total = raw.get("total_anchors", 0)
    people_raw = raw.get("people", [])
    mutual_mentions = raw.get("mutual_mentions", 0)
    console.print(
        f"[dim]  (scan: {total} profile links on page, {len(people_raw)} unique, "
        f"{mutual_mentions} 'mutual connection' mentions)[/dim]"
    )

    # If the anchor-based pass found people, use it. Otherwise fall back to a
    # text-driven pass: find every element whose text mentions a mutual
    # connection and grab the nearest /in/ link within it. LinkedIn sometimes
    # structures PYMK cards so the profile link isn't the obvious anchor.
    if not people_raw:
        fallback = await page.evaluate(
            """() => {
                const out = [];
                const seen = new Set();
                const all = Array.from(document.querySelectorAll('div, li, section, article'));
                for (const el of all) {
                    const t = (el.innerText || '').trim();
                    if (!/mutual connection/i.test(t)) continue;
                    if (t.length > 600) continue;  // too big = not a single card
                    const link = el.querySelector("a[href*='/in/']");
                    let href = link ? (link.getAttribute('href') || '') : '';
                    try { if (href) href = new URL(href, location.origin).pathname; } catch (e) {}
                    const key = href || t.slice(0, 40);
                    if (seen.has(key)) continue;
                    seen.add(key);
                    out.push({ href, text: t });
                }
                return out;
            }"""
        )
        if fallback:
            console.print(f"[dim]  (text-fallback found {len(fallback)} cards mentioning mutual connections)[/dim]")
            people_raw = fallback

    # If we STILL found nothing usable, dump everything so we can see why —
    # but only on the final retry (caller passes dump_on_empty), so we don't
    # spam dumps for intermediate attempts while cards are still loading.
    if not people_raw and dump_on_empty:
        console.print("[yellow]  No usable profile cards in the DOM after retries. "
                      "Saving a debug dump so we can see what the page actually contained.[/yellow]")
        await dump_debug(page, "no_links")

    hints = []
    for item in people_raw:
        href = (item.get("href") or "").strip()
        text = (item.get("text") or "").strip()

        if href and "/in/" in href:
            profile_url = "https://www.linkedin.com" + href if href.startswith("/") else href
        else:
            profile_url = None  # text-fallback card with no captured link

        # Skip cards that are neither a person link nor mention mutuals
        if profile_url is None and "mutual" not in text.lower():
            continue

        # Prefer the per-anchor name (reliable for THIS person). Only fall
        # back to scanning the block's lines when the anchor had no text.
        anchor_name = (item.get("anchor_name") or "").strip()
        anchor_name = re.sub(r'\s*·\s*(1st|2nd|3rd).*$', '', anchor_name).strip()

        name = anchor_name
        if not name:
            for line in text.split("\n"):
                line = line.strip()
                low = line.lower()
                if (2 <= len(line) <= 60 and "mutual" not in low and "degree" not in low
                        and low not in ("connect", "message", "follow", "following",
                                        "pending", "1st", "2nd", "3rd", "· 1st", "· 2nd", "· 3rd")):
                    name = re.sub(r'\s*·\s*(1st|2nd|3rd)\s*$', '', line).strip()
                    break

        # Drop service-provider blurbs and other non-person rows entirely —
        # these are not people and must not appear as candidates.
        if is_non_person_name(name):
            continue

        hints.append({
            "name": name or "(name unknown — see card)",
            "mutual_count": parse_mutual_count(text),
            "degree": parse_degree(text),
            "profile_url": profile_url,
        })

    # Dedupe: LinkedIn often exposes the same person via two URL forms (a
    # vanity '/in/name' link and an opaque '/in/ACoAA...' link), producing
    # duplicate rows. Collapse by display name, preferring the entry that has
    # a real name, a known mutual count, and a vanity (non-ACoAA) URL.
    def _is_vanity(u):
        return bool(u) and "/in/ACoAA" not in u

    deduped = {}
    order = []
    for h in hints:
        key = h["name"].strip().lower()
        if not key or key.startswith("(name unknown"):
            # Keep unnamed entries as-is (can't safely merge them)
            order.append(h)
            continue
        if key not in deduped:
            deduped[key] = h
            order.append(h)
        else:
            cur = deduped[key]
            # Prefer a known count over None
            if cur["mutual_count"] is None and h["mutual_count"] is not None:
                cur["mutual_count"] = h["mutual_count"]
            # Prefer a vanity URL over an opaque one
            if not _is_vanity(cur.get("profile_url")) and _is_vanity(h.get("profile_url")):
                cur["profile_url"] = h["profile_url"]
            # Prefer a known degree
            if cur.get("degree") in (None, "unknown") and h.get("degree") not in (None, "unknown"):
                cur["degree"] = h["degree"]

    # Rebuild preserving original order, skipping merged duplicates
    final = []
    emitted = set()
    for h in order:
        key = h["name"].strip().lower()
        if key and not key.startswith("(name unknown"):
            if key in emitted:
                continue
            emitted.add(key)
            final.append(deduped[key])
        else:
            final.append(h)

    return final


def print_hints_table(company: str, hints: list[dict]):
    table = Table(
        title=f"👀 Visible People-Tab Hints — {company}",
        show_header=True,
        header_style="bold cyan",
        border_style="dim",
    )
    table.add_column("Name", style="bold white")
    table.add_column("Degree", style="cyan")
    table.add_column("Mutuals", style="yellow")
    table.add_column("Flag", style="bold red")

    flagged_count = 0
    for h in hints:
        is_flagged = h["mutual_count"] is not None and h["mutual_count"] >= MIN_MUTUAL_CONNECTIONS
        if is_flagged:
            flagged_count += 1
        table.add_row(
            h["name"],
            h.get("degree", "—"),
            str(h["mutual_count"]) if h["mutual_count"] is not None else "—",
            f"🔥 {MIN_MUTUAL_CONNECTIONS}+" if is_flagged else "",
        )

    console.print(table)
    console.print(
        f"[dim]{flagged_count} of {len(hints)} visible people already meet the "
        f"{MIN_MUTUAL_CONNECTIONS}+ mutuals bar (per the list view — re-checked on click). "
        f"For everyone else, you'll get a clickable list of their mutuals to check the "
        f"<10-but-influential-mutual case during one-by-one review.[/dim]\n"
    )

    # Dedicated summary: real candidates are 1st/2nd-degree people who
    # actually share at least one mutual connection. An unknown/None count
    # almost always means zero mutuals (the card simply had no mutual line),
    # and those rows are often non-people junk ("Provides services...",
    # "Andy works here"), so we exclude them rather than list them.
    candidates = [
        h for h in hints
        if h.get("degree") in ("1st", "2nd")
        and h.get("mutual_count") is not None
        and h.get("mutual_count") >= 1
    ]
    if candidates:
        console.print(f"[bold green]🟢 1st / 2nd-degree candidates with shared mutuals for {company} "
                      f"({len(candidates)}):[/bold green]")
        for h in candidates:
            deg = h.get("degree", "unknown")
            mc = h["mutual_count"]
            flag = "  🔥 10+" if mc >= MIN_MUTUAL_CONNECTIONS else ""
            url = h.get("profile_url") or ""
            console.print(f"  • [white]{h['name']}[/white] [cyan]({deg})[/cyan] — {mc} mutuals{flag}")
            if url:
                console.print(f"    [dim]{url}[/dim]")
        console.print()
    else:
        console.print(f"[yellow]No 1st/2nd-degree candidates with shared mutuals detected among the "
                      f"loaded cards for {company}.[/yellow]\n")


async def get_profile_mutual_info(page: Page, profile_url: str, harvest_mutuals: bool = True,
                                  authoritative_count: Optional[int] = None) -> dict:
    """Open a profile and extract degree + mutual-connection count, using
    plain visible text as the PRIMARY source (most resistant to LinkedIn
    changing its CSS class names).

    authoritative_count, when provided (the count shown on the People-tab
    card), is treated as the trustworthy number of shared mutuals and used to
    cap the harvested list — this prevents the 'profile text says 8 but the
    card said 4' over-reporting, since the profile body text can contain
    stray 'N mutual connections' strings from other UI sections."""

    await safe_goto(page, profile_url)
    await human_delay(2, 4)
    await scroll_page(page, 2)

    degree = "unknown"
    mutual_count = 0
    mutuals = []  # [{"name": str, "profile_url": str}]

    try:
        degree_badge = await page.query_selector(
            ".dist-value, span[class*='distance-badge'], .artdeco-entity-lockup__badge"
        )
        if degree_badge:
            degree = (await degree_badge.text_content() or "").strip()

        body_text = await page.evaluate("document.body.innerText")
        parsed = parse_mutual_count(body_text)
        if parsed is not None:
            mutual_count = parsed

        # The People-tab card count is more reliable than scraping the whole
        # profile body (which can include unrelated 'mutual connections' text).
        if authoritative_count is not None:
            mutual_count = authoritative_count

        # Harvest the mutual list for anyone who shares at least one mutual.
        need_list = harvest_mutuals and mutual_count >= 1
        if need_list:
            cap = authoritative_count if authoritative_count is not None else mutual_count
            mutuals = await harvest_mutual_connections(page, profile_url, cap)
            if await current_url(page) != profile_url:
                await safe_goto(page, profile_url)
                await human_delay(1.0, 2.0)

        if mutual_count == 0:
            match3 = re.search(r'You both know (.+?)(?:\n|$)', body_text)
            if match3:
                mutual_count = 1
                if not mutuals:
                    mutuals = [{"name": match3.group(1).strip(), "profile_url": None}]

    except Exception as e:
        console.print(f"[dim red]  ⚠ Error parsing mutuals: {e}[/dim red]")

    return {"mutual_count": mutual_count, "mutuals": mutuals, "degree": degree}


async def harvest_mutual_connections(page: Page, profile_url: str, expected_count: int) -> list[dict]:
    """Fetch the REAL list of shared mutual connections for the profile that
    is CURRENTLY open (`page` must already be on `profile_url`).

    Strategy: read the href of the profile's own 'X mutual connections' link
    (it points at the correct mutual-connections search for THIS member) and
    navigate to it directly — more reliable than clicking, which sometimes
    opens an overlay that doesn't change the URL. Then scrape /in/ links from
    that results page. Cap to expected_count so we never over-report.
    """
    mutuals = []
    if expected_count <= 0:
        return mutuals

    try:
        owner_id_m = re.search(r'/in/([^/?#]+)', profile_url)
        owner_id = owner_id_m.group(1) if owner_id_m else ""

        # Grab the href of the profile's own mutual-connections link.
        mutual_href = await page.evaluate(
            """() => {
                const links = Array.from(document.querySelectorAll('a'));
                for (const a of links) {
                    const t = (a.innerText || '').toLowerCase();
                    const h = a.getAttribute('href') || '';
                    if (/mutual connection/.test(t) &&
                        (h.includes('/search/results/people') ||
                         h.includes('connectionOf') || h.includes('facetNetwork'))) {
                        return h;
                    }
                }
                // Fallback: any link whose text mentions mutual connections
                for (const a of links) {
                    if (/mutual connection/i.test(a.innerText || '')) {
                        return a.getAttribute('href') || '';
                    }
                }
                return '';
            }"""
        )

        scraped = []
        if mutual_href:
            full_mutual_url = mutual_href
            if mutual_href.startswith("/"):
                full_mutual_url = "https://www.linkedin.com" + mutual_href
            await safe_goto(page, full_mutual_url)
            await human_delay(2, 3)
            await thorough_scroll(page, max_steps=4)
            scraped = await _scrape_people_links(page, owner_id)

        # Fallback: if the link approach gave nothing, scrape the inline
        # mutual section straight off the profile (re-open it first).
        if not scraped:
            await safe_goto(page, profile_url)
            await human_delay(1.5, 2.5)
            await scroll_page(page, 2)
            scraped = await _scrape_people_links(page, owner_id, inline_only=True)

        # Cap strictly to the expected (authoritative) count so we never list
        # more mutuals than the person actually shares with you.
        if expected_count > 0:
            scraped = scraped[:expected_count]
        mutuals = scraped[:MAX_MUTUALS_PER_PERSON]

    except Exception as e:
        console.print(f"[dim red]  ⚠ Couldn't harvest mutual list: {e}[/dim red]")

    return mutuals


async def _scrape_people_links(page: Page, owner_id: str, inline_only: bool = False) -> list[dict]:
    """Scrape (name, profile_url) for /in/ links on the current page, skipping
    the owner and obvious UI/sidebar strings. If inline_only, restrict to a
    container near a 'mutual connections' heading on a profile page."""
    raw = await page.evaluate(
        """(args) => {
            const ownerId = args.ownerId;
            const inlineOnly = args.inlineOnly;

            // Headings that mark RECOMMENDATION sections (not mutuals). Any
            // /in/ link under one of these is noise.
            const NOISE_HEADINGS = [
                'more profiles for you', 'people you may know', 'you might like',
                'pages for you', 'ai agents', 'products you might', 'from ',
                'similar profiles', 'also viewed', 'also follow', 'promoted'
            ];
            const inNoiseSection = (el) => {
                let n = el;
                for (let i = 0; i < 14 && n; i++) {
                    // Check preceding heading-ish siblings' text
                    let sib = n.previousElementSibling;
                    let hops = 0;
                    while (sib && hops < 4) {
                        const t = (sib.innerText || '').toLowerCase().slice(0, 60);
                        if (NOISE_HEADINGS.some(h => t.includes(h))) return true;
                        sib = sib.previousElementSibling; hops++;
                    }
                    n = n.parentElement;
                }
                return false;
            };

            let scope = document;
            if (inlineOnly) {
                const all = Array.from(document.querySelectorAll('section, div'));
                for (const el of all) {
                    if (/mutual connection/i.test(el.innerText || '') &&
                        (el.innerText || '').length < 1200) { scope = el; break; }
                }
            }
            const out = [];
            const seen = new Set();
            const anchors = Array.from(scope.querySelectorAll("a[href*='/in/']"));
            for (const a of anchors) {
                if (inNoiseSection(a)) continue;
                let href = a.getAttribute('href') || '';
                try { href = new URL(href, location.origin).pathname; } catch (e) {}
                if (!href.includes('/in/')) continue;
                if (ownerId && href.includes('/in/' + ownerId)) continue;
                if (seen.has(href)) continue;
                seen.add(href);
                let node = a;
                let txt = (a.innerText || '').trim();
                for (let i = 0; i < 4 && node && !txt; i++) {
                    node = node.parentElement;
                    if (node) txt = (node.innerText || '').trim();
                }
                const name = (txt.split('\\n')[0] || '').trim();
                out.push({ href, name });
            }
            return out;
        }""",
        {"ownerId": owner_id, "inlineOnly": inline_only},
    )

    results = []
    for item in raw:
        href = item["href"]
        name = re.sub(r'\s*·\s*(1st|2nd|3rd).*$', '', item.get("name", "")).strip()
        low = name.lower()
        if not name or len(name) > 60:
            continue
        if any(k in low for k in ("show all", "see all", "results", "follower",
                                  "course", "language", "promoted", "follow")):
            continue
        full = "https://www.linkedin.com" + href if href.startswith("/") else href
        results.append({"name": name, "profile_url": full})
    return results


async def check_specific_profile_mutuals(page: Page, profile_url: str, return_to_url: Optional[str] = None) -> int:
    """Given a direct profile URL (e.g. a mutual you clicked on), return how
    many mutual connections YOU have with THEM — this is the exact metric
    the second-degree criterion needs (a mutual who has 20+ mutuals with
    you), not their total connection count."""
    count = 0
    try:
        info = await get_profile_mutual_info(page, profile_url, harvest_mutuals=False)
        count = info["mutual_count"]
    except Exception as e:
        console.print(f"[red]⚠ Error checking profile: {e}[/red]")
    finally:
        if return_to_url and await current_url(page) != return_to_url:
            await safe_goto(page, return_to_url)
            await human_delay(1.5, 2.5)
    return count


async def check_mutual_by_name(page: Page, mutual_name: str, return_to_url: Optional[str] = None) -> int:
    """Fallback manual check when a mutual's profile link wasn't captured:
    search their name, open the top result, and read mutual-connection
    count using the same (corrected) metric as check_specific_profile_mutuals."""
    count = 0
    try:
        search_url = f"https://www.linkedin.com/search/results/people/?keywords={mutual_name.replace(' ', '%20')}"
        await safe_goto(page, search_url)
        await human_delay(2, 4)

        first_result = await page.query_selector("a[href*='/in/']")
        if not first_result:
            return 0
        href = await first_result.get_attribute("href")
        if not href:
            return 0

        info = await get_profile_mutual_info(page, href.split("?")[0], harvest_mutuals=False)
        count = info["mutual_count"]
    except Exception as e:
        console.print(f"[red]⚠ Error checking '{mutual_name}': {e}[/red]")
    finally:
        if return_to_url and await current_url(page) != return_to_url:
            await safe_goto(page, return_to_url)
            await human_delay(1.5, 2.5)
    return count


# ─────────────────────────────────────────────
# AUTOMATED PROCESSING (no human interaction)
# ─────────────────────────────────────────────

async def process_person(page: Page, person: dict) -> dict:
    """Open one person's profile and classify them by how many mutual
    connections they share WITH YOU (this is what 'degree' means here, NOT
    LinkedIn's network distance):

      - 10+ shared mutuals  -> '1st-degree lead'. Recorded as a qualifying
        lead. We do NOT list/open their mutuals (too slow, not needed).
      - 1..9 shared mutuals -> '2nd-degree lead'. We harvest the actual
        mutual list and open each mutual to see whether any of THEM shares
        20+ mutuals with you; if so, this person qualifies too.
      - 0 shared mutuals    -> not a lead.
    """
    name = person.get("name", "")
    profile_url = person.get("profile_url", "")
    list_count = person.get("mutual_count")  # count seen on the People-tab card
    console.print(f"    → {name} ...")

    record = {
        "name": name,
        "profile_url": profile_url,
        "degree": "",            # '1st-degree lead' / '2nd-degree lead'
        "mutual_count": list_count,
        "category": "",
        "qualifies": False,
        "qualify_reason": "",
        "mutuals": [],           # list of {name, profile_url, their_mutuals_with_me}
    }

    # Decide first whether this is a 10+ (1st-degree) lead. For those we skip
    # the expensive mutual harvest entirely. We rely on the list-view count;
    # only when it's under 10 do we open the profile to get the true list.
    is_first_degree = list_count is not None and list_count >= MIN_MUTUAL_CONNECTIONS

    if is_first_degree:
        record["degree"] = "1st-degree lead"
        record["mutual_count"] = list_count
        record["category"] = f"{MIN_MUTUAL_CONNECTIONS}+ mutuals"
        record["qualifies"] = True
        record["qualify_reason"] = f"{list_count} mutual connections (>= {MIN_MUTUAL_CONNECTIONS})"
        console.print(f"      {list_count} mutuals → 1st-degree lead  ✅ (mutuals not listed)")
        return record

    # Under 10 (or unknown): open profile, harvest the real mutual list.
    try:
        info = await get_profile_mutual_info(page, profile_url, harvest_mutuals=True,
                                             authoritative_count=list_count)
    except Exception as e:
        msg = str(e)
        if any(k in msg for k in ("Execution context was destroyed", "Target closed",
                                  "navigation", "detached")):
            console.print(f"      [yellow](page navigated; retrying {name} once)[/yellow]")
            try:
                info = await get_profile_mutual_info(page, profile_url, harvest_mutuals=True,
                                                     authoritative_count=list_count)
            except Exception:
                info = {"mutual_count": list_count or 0, "mutuals": [], "degree": ""}
        else:
            console.print(f"      [red]error reading {name}: {e}[/red]")
            info = {"mutual_count": list_count or 0, "mutuals": [], "degree": ""}

    harvested = info.get("mutuals", []) or []

    # If the profile read now shows 10+, treat as a 1st-degree lead after all
    # (and don't bother enumerating mutuals).
    profile_count = info.get("mutual_count")
    if profile_count is not None and profile_count >= MIN_MUTUAL_CONNECTIONS:
        record["degree"] = "1st-degree lead"
        record["mutual_count"] = profile_count
        record["category"] = f"{MIN_MUTUAL_CONNECTIONS}+ mutuals"
        record["qualifies"] = True
        record["qualify_reason"] = f"{profile_count} mutual connections (>= {MIN_MUTUAL_CONNECTIONS})"
        console.print(f"      {profile_count} mutuals → 1st-degree lead  ✅ (mutuals not listed)")
        return record

    # The mutual count: prefer the authoritative People-tab card count
    # (what you saw on the list). The harvested list is what we enumerate and
    # check, but the count shown should match the card to avoid the
    # 'shows 8 but has 4' mismatch.
    if list_count is not None and list_count >= 1:
        mutual_count = list_count
    elif harvested:
        mutual_count = len(harvested)
    else:
        mutual_count = profile_count if profile_count is not None else 0

    record["mutual_count"] = mutual_count
    record["degree"] = "2nd-degree lead" if mutual_count >= 1 else "not a lead"

    # 2nd-degree: open each mutual to see if any shares 20+ mutuals with you.
    influential_found = False
    enriched = []
    for m in harvested:
        their_count = None
        if m.get("profile_url"):
            try:
                their_count = await check_specific_profile_mutuals(
                    page, m["profile_url"], return_to_url=profile_url
                )
            except Exception:
                their_count = None
            await human_delay(*DELAY_BETWEEN_MUTUALS)
        if their_count is not None and their_count >= MIN_MUTUAL_FRIEND_CONNECTIONS:
            influential_found = True
        enriched.append({
            "name": m.get("name", ""),
            "profile_url": m.get("profile_url", ""),
            "their_mutuals_with_me": their_count,
        })
    record["mutuals"] = enriched

    # Classify the under-10 person
    if mutual_count >= 1 and influential_found:
        record["category"] = f"<{MIN_MUTUAL_CONNECTIONS} but strong mutual"
        record["qualifies"] = True
        strong = [m for m in enriched if (m["their_mutuals_with_me"] or 0) >= MIN_MUTUAL_FRIEND_CONNECTIONS]
        names = ", ".join(m["name"] for m in strong[:3])
        record["qualify_reason"] = (
            f"{mutual_count} mutuals; influential mutual(s) sharing "
            f"{MIN_MUTUAL_FRIEND_CONNECTIONS}+ with you: {names}"
        )
    elif mutual_count >= 1:
        record["category"] = f"<{MIN_MUTUAL_CONNECTIONS} mutuals"
        record["qualifies"] = False
        record["qualify_reason"] = (
            f"{mutual_count} mutuals; none shares {MIN_MUTUAL_FRIEND_CONNECTIONS}+ with you"
        )
    else:
        record["category"] = "no shared mutuals"
        record["qualifies"] = False
        record["qualify_reason"] = "0 mutual connections"

    flag = "✅ QUALIFIES" if record["qualifies"] else ""
    console.print(f"      {mutual_count} mutuals → {record['degree']}, "
                  f"{len(enriched)} checked  {flag}")
    return record


async def process_company(cpage: Page, people_url: str, company_label: str) -> dict:
    """Fully automated processing of one company: load the People tab, scan
    everyone, then process each 1st/2nd-degree person with shared mutuals."""
    # Fresh navigation forces LinkedIn to render the People list (first-try fix)
    await safe_goto(cpage, people_url)
    await human_delay(*DELAY_AFTER_NAV)
    await scroll_page(cpage, 2)
    await click_show_more(cpage, max_clicks=4)

    company_name = await get_company_display_name(cpage, company_label)
    console.print(f"\n[bold]🏢 {company_name}[/bold]  [dim]{people_url}[/dim]")

    hints = await scan_visible_people_hints(cpage, reload_url=people_url)
    print_hints_table(company_name, hints)

    # Candidates: 1st/2nd-degree with at least one shared mutual
    candidates = [
        h for h in hints
        if h.get("profile_url")
        and h.get("degree") in ("1st", "2nd")
        and h.get("mutual_count") is not None
        and h.get("mutual_count") >= 1
        and not is_non_person_name(h.get("name", ""))
    ]
    console.print(f"[dim]Processing {len(candidates)} qualifying-degree people for {company_name}...[/dim]")

    records = []
    for person in candidates:
        rec = await process_person(cpage, person)
        rec["company"] = company_name
        rec["company_url"] = people_url
        records.append(rec)
        await human_delay(*DELAY_BETWEEN_PROFILES)

    return {"company": company_name, "people_url": people_url, "records": records}


def export_to_excel(all_results: list[dict], out_path: Path) -> Path:
    """Write all collected data to a formatted .xlsx with two sheets:
    a 'People' summary and a 'Mutuals' detail sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    header_fill = PatternFill("solid", start_color="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    qualify_fill = PatternFill("solid", start_color="C6EFCE")
    base_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.freeze_panes = "A2"

    # ── Sheet 1: People summary ──
    ws1 = wb.active
    ws1.title = "People"
    headers1 = ["Company", "Name", "Lead type", "Mutual count", "Category",
                "Qualifies", "Why", "# Mutuals listed",
                "# Influential mutuals (20+)", "Profile URL"]
    ws1.append(headers1)

    for result in all_results:
        for r in result["records"]:
            influ = sum(1 for m in r["mutuals"]
                        if (m.get("their_mutuals_with_me") or 0) >= MIN_MUTUAL_FRIEND_CONNECTIONS)
            ws1.append([
                r.get("company", ""),
                r.get("name", ""),
                r.get("degree", ""),
                r.get("mutual_count", 0),
                r.get("category", ""),
                "YES" if r.get("qualifies") else "no",
                r.get("qualify_reason", ""),
                len(r.get("mutuals", [])),
                influ,
                r.get("profile_url", ""),
            ])

    style_header(ws1, len(headers1))
    widths1 = [22, 26, 8, 12, 22, 10, 46, 14, 16, 50]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    for row in range(2, ws1.max_row + 1):
        for col in range(1, len(headers1) + 1):
            cell = ws1.cell(row=row, column=col)
            cell.font = base_font
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(col in (5, 7)))
        if ws1.cell(row=row, column=6).value == "YES":
            for col in range(1, len(headers1) + 1):
                ws1.cell(row=row, column=col).fill = qualify_fill

    # ── Sheet 2: Mutuals detail ──
    ws2 = wb.create_sheet("Mutuals")
    headers2 = ["Company", "Person", "Person lead type", "Person mutual count",
                "Mutual name", "Mutual's shared-with-me count",
                "Influential (20+)?", "Mutual profile URL"]
    ws2.append(headers2)

    for result in all_results:
        for r in result["records"]:
            # If there are no shared mutuals, don't list the person on this
            # sheet at all (they still appear on the People summary sheet).
            if not r["mutuals"]:
                continue
            for m in r["mutuals"]:
                tc = m.get("their_mutuals_with_me")
                ws2.append([
                    r.get("company", ""),
                    r.get("name", ""),
                    r.get("degree", ""),
                    r.get("mutual_count", 0),
                    m.get("name", ""),
                    tc if tc is not None else "",
                    "YES" if (tc or 0) >= MIN_MUTUAL_FRIEND_CONNECTIONS else "",
                    m.get("profile_url", ""),
                ])

    style_header(ws2, len(headers2))
    widths2 = [22, 24, 12, 16, 26, 22, 14, 50]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for row in range(2, ws2.max_row + 1):
        for col in range(1, len(headers2) + 1):
            cell = ws2.cell(row=row, column=col)
            cell.font = base_font
            cell.border = border
            cell.alignment = Alignment(vertical="top")
        if ws2.cell(row=row, column=7).value == "YES":
            for col in range(1, len(headers2) + 1):
                ws2.cell(row=row, column=col).fill = qualify_fill

    # Write atomically: save to a temp file, then replace the target. If the
    # target is locked (you have it open in Excel), fall back to a timestamped
    # filename so the run never crashes mid-way.
    out_path = Path(out_path)
    tmp_path = out_path.with_suffix(".tmp.xlsx")
    wb.save(tmp_path)
    try:
        import os
        os.replace(tmp_path, out_path)
        return out_path
    except Exception:
        alt = out_path.with_name(
            f"{out_path.stem}_{datetime.now().strftime('%H%M%S')}{out_path.suffix}"
        )
        try:
            os.replace(tmp_path, alt)
        except Exception:
            wb.save(alt)
        console.print(f"[yellow]  (target file was open/locked — wrote {alt.name} instead)[/yellow]")
        return alt


def _looks_logged_in(url: str) -> bool:
    """True if the URL is a real logged-in LinkedIn page (feed / my network).
    Checks the PATH only — a logged-out visit to /feed/ redirects to
    /uas/login?session_redirect=...%2Ffeed%2F, which still contains 'feed' in
    its query string, so a naive substring test would be fooled into thinking
    we're logged in (and could even save a logged-out session)."""
    base = (url or "").split("?")[0].split("#")[0].lower()
    return "/feed" in base or "/mynetwork" in base


async def run_login(timeout_s: int = 300) -> bool:
    """One-time interactive login. Opens a browser, lets the user sign in to
    LinkedIn (including any 2FA / security check), then AUTO-DETECTS the logged-
    in feed and saves the session — no terminal input required. Run once before
    using the scanner through an agent; the saved session is reused after that."""
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=False,  # the user must see the page to log in
            args=["--disable-blink-features=AutomationControlled"],
        )
        ctx_kwargs = dict(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 900},
        )
        if SESSION_FILE.exists():
            context = await browser.new_context(storage_state=str(SESSION_FILE), **ctx_kwargs)
        else:
            context = await browser.new_context(**ctx_kwargs)
        page = await context.new_page()
        await page.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            window.chrome = {runtime: {}};
        """)

        await safe_goto(page, "https://www.linkedin.com/feed/")
        await human_delay(2, 3)
        url_now = await current_url(page)
        if _looks_logged_in(url_now):
            console.print("[green]✓ Already logged in — saving session.[/green]")
            await save_session(context)
            await browser.close()
            return True

        await safe_goto(page, "https://www.linkedin.com/login")
        console.print(Panel(
            "A browser window has opened.\n"
            "1. Log in to LinkedIn (finish any 2FA / security check).\n"
            "2. Wait until your LinkedIn home feed appears.\n\n"
            "This tool detects it automatically and saves your session — you do "
            "NOT need to come back to the terminal or press anything.",
            title="🔐 Log in to LinkedIn (one time)",
            border_style="yellow",
        ))

        # Poll for the logged-in state rather than blocking on input(), so this
        # works whether a person or an agent kicked it off.
        deadline = time.time() + timeout_s
        while time.time() < deadline:
            await asyncio.sleep(3)
            url_now = await current_url(page)
            if _looks_logged_in(url_now):
                await human_delay(1, 2)
                await save_session(context)
                console.print("[green]✓ Login detected — session saved. You can close the browser.[/green]")
                await browser.close()
                return True

        console.print(
            f"[red]Timed out after {timeout_s}s waiting for login. "
            f"Re-run [bold]python main.py --login[/bold] and finish signing in.[/red]"
        )
        await browser.close()
        return False


async def run_scout_automated(headless: bool = False, urls: list[str] = None,
                              reset_progress: bool = False):
    """Fully automated, URL-driven run: no manual picking, no pauses. Processes
    every company URL, collects data, and exports an Excel file at the end."""
    company_urls = urls or COMPANY_URLS
    people_urls = []
    for u in company_urls:
        pu = normalize_company_people_url(u)
        if pu:
            people_urls.append(pu)
        else:
            console.print(f"[yellow]Skipping unrecognized URL: {u}[/yellow]")

    if reset_progress:
        if PROGRESS_FILE.exists():
            PROGRESS_FILE.unlink()
        if RESULTS_FILE.exists():
            RESULTS_FILE.unlink()
    done = load_progress()

    console.print("\n[bold cyan]━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━[/bold cyan]")
    console.print("[bold cyan]  🔍 LinkedIn Scout — Automated Run [/bold cyan]")
    console.print("[bold cyan]━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━[/bold cyan]")
    console.print(f"[dim]{len(people_urls)} companies queued. Output → {OUTPUT_XLSX}[/dim]\n")

    if not people_urls:
        console.print("[red]No valid company URLs in COMPANY_URLS. Paste some and re-run.[/red]")
        return

    # Seed with previously-scraped companies so skipped ones stay in the sheet.
    all_results = load_results()

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=headless,
            args=["--disable-blink-features=AutomationControlled"],
        )
        ctx_kwargs = dict(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 900},
        )
        if SESSION_FILE.exists():
            console.print(f"[green]✓ Loading saved session from {SESSION_FILE}[/green]")
            context = await browser.new_context(storage_state=str(SESSION_FILE), **ctx_kwargs)
        else:
            context = await browser.new_context(**ctx_kwargs)

        page = await context.new_page()
        await page.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            window.chrome = {runtime: {}};
        """)

        await safe_goto(page, "https://www.linkedin.com/feed/")
        await human_delay(2, 3)

        # Scan runs are NON-interactive (an agent can't answer a prompt), so we
        # never pause for login here. If there's no valid session, stop with a
        # clear message telling the user to run `--login` once.
        url_now = await current_url(page)
        if not _looks_logged_in(url_now):
            console.print(
                "\n[bold red]🚫 Not logged in to LinkedIn (no valid session).[/bold red]\n"
                "[yellow]Run this once to sign in, then re-run the scan:[/yellow]\n"
                "    [bold]python main.py --login[/bold]\n"
            )
            await browser.close()
            return

        try:
            for people_url in people_urls:
                if people_url in done:
                    console.print(f"[dim]Already done: {people_url} — skipping[/dim]")
                    continue

                label = company_name_from_url(people_url)
                console.print(f"\n[bold]📌 {label}[/bold]")

                # Block check; if walled, save what we have and stop politely.
                block = await detect_block(page)
                if block:
                    console.print(f"[bold red]🚫 {block} — pausing run to protect the account.[/bold red]")
                    break

                try:
                    result = await process_company(page, people_url, label)
                    # Replace any stale entry for this company, then append, so
                    # re-processing never duplicates a row in the spreadsheet.
                    all_results = [r for r in all_results
                                   if r.get("people_url") != people_url]
                    all_results.append(result)
                    save_results(all_results)
                    done.add(people_url)
                    save_progress(done)

                    # Regenerate the spreadsheet after EACH company so you can
                    # open it and verify the data as the run progresses,
                    # instead of waiting until the very end.
                    try:
                        out = export_to_excel(all_results, OUTPUT_XLSX)
                        abs_path = Path(out).resolve()
                        file_link = abs_path.as_uri()  # file:///C:/... openable link
                        n_people = len(result["records"])
                        n_qual = sum(1 for p in result["records"] if p["qualifies"])
                        console.print(
                            f"[green]  💾 Spreadsheet updated ({label}: {n_people} people, "
                            f"{n_qual} qualifying)[/green]"
                        )
                        console.print(f"     [bold cyan]Open it here:[/bold cyan] {file_link}")
                        console.print(f"     [dim]{abs_path}[/dim]")
                    except Exception as ex:
                        console.print(f"[yellow]  ⚠ Couldn't update spreadsheet after {label}: {ex}[/yellow]")
                except Exception as e:
                    console.print(f"[red]Error on {label}: {e} — moving on.[/red]")

                await human_delay(*DELAY_BETWEEN_COMPANIES)
        finally:
            await browser.close()

    # Export whatever we collected (even if interrupted)
    if all_results:
        out = export_to_excel(all_results, OUTPUT_XLSX)
        total_people = sum(len(r["records"]) for r in all_results)
        qualify = sum(1 for r in all_results for p in r["records"] if p["qualifies"])
        console.print(f"\n[bold green]✓ Done. {len(all_results)} companies, {total_people} people "
                      f"({qualify} qualifying). Excel saved → {out}[/bold green]")
    else:
        console.print("\n[yellow]No data collected — nothing to export.[/yellow]")


# ─────────────────────────────────────────────
# CLI Entry Point
# ─────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="LinkedIn Scout - Automated company People-tab scraper -> Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Setup:
  1. pip install -r requirements.txt  &&  playwright install chromium
  2. python main.py --login           (one time: opens browser, sign in, auto-saves)
  3. python main.py <company urls...>  (scan; pass URLs as arguments)
  4. When it finishes, open linkedin_scout_results.xlsx.

Examples:
  python main.py https://www.linkedin.com/company/acme/ https://www.linkedin.com/company/foo/
  python main.py --urls-file companies.txt   # one company URL per line
  python main.py                  # Fall back to COMPANY_URLS in this file
  python main.py --headless       # Run in background
  python main.py --reset-progress # Re-process all companies from scratch
  python main.py --reset-session  # Force re-login

URL input (for agent / non-technical use):
  Pass company URLs straight on the command line, or put them (one per line) in a
  text file and pass --urls-file. Either overrides the COMPANY_URLS list in this
  file, so callers never have to edit the source. Lines starting with # are
  ignored in a urls file.
        """
    )
    parser.add_argument("urls", nargs="*", help="Company LinkedIn URLs to scan (overrides COMPANY_URLS)")
    parser.add_argument("--urls-file", help="Path to a text file of company URLs, one per line")
    parser.add_argument("--login", action="store_true",
                        help="One-time: open a browser to log in to LinkedIn and save the session")
    parser.add_argument("--headless", action="store_true", help="Run browser in headless mode")
    parser.add_argument("--reset-session", action="store_true", help="Delete saved session and re-login")
    parser.add_argument("--reset-progress", action="store_true", help="Re-process all companies")

    args = parser.parse_args()

    if args.reset_session and SESSION_FILE.exists():
        SESSION_FILE.unlink()
        console.print("[yellow]Session cleared. Will re-login.[/yellow]")

    # One-time login mode: sign in, save session, exit. Everything else (the
    # actual scans) then runs unattended using that saved session.
    if args.login:
        ok = asyncio.run(run_login())
        sys.exit(0 if ok else 1)

    # Collect URLs from the command line and/or a file; fall back to the
    # hard-coded COMPANY_URLS only when the caller supplied none. This is what
    # lets an agent pass a user's URLs without anyone editing the source.
    cli_urls = list(args.urls)
    if args.urls_file:
        try:
            for line in Path(args.urls_file).read_text(encoding="utf-8").splitlines():
                line = line.strip()
                if line and not line.startswith("#"):
                    cli_urls.append(line)
        except Exception as e:
            console.print(f"[red]Could not read --urls-file {args.urls_file}: {e}[/red]")

    asyncio.run(run_scout_automated(
        headless=args.headless,
        urls=cli_urls or None,
        reset_progress=args.reset_progress,
    ))
